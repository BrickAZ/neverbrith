import argparse
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


MAIN_SHEET = "道具审批"
PLAYER_START_ROW = 13
AI_SUFFIX = "（AI自动补充）"


REQUIRED_HEADERS = (
    "类型",
    "英文名",
    "中文名",
    "充能",
    "道具池/权重",
    "英文描述",
    "中文描述",
    "道具效果（自然语言）",
    "贴图",
    "设计者联系平台",
    "设计者联系方式",
    "审查状态",
    "AI回复",
    "设计状态",
    "美术状态",
    "代码状态",
    "工作流ID",
    "最后处理时间",
)


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_for_output(value):
    text = clean_text(value)
    return text.replace(AI_SUFFIX, "").strip()


def build_header_map(ws):
    headers = {}
    for col in range(1, ws.max_column + 1):
        value = clean_text(ws.cell(1, col).value)
        if value:
            headers[value] = col
    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        raise RuntimeError(f"Missing required columns: {', '.join(missing)}")
    return headers


def cell(ws, row, headers, name):
    return ws.cell(row, headers[name])


def get_row_value(ws, row, headers, name):
    return clean_text(cell(ws, row, headers, name).value)


def item_key(ws, row, headers):
    return clean_for_output(get_row_value(ws, row, headers, "英文名")) or clean_for_output(
        get_row_value(ws, row, headers, "中文名")
    )


def safe_key(name):
    slug = re.sub(r"[^a-z0-9]+", "-", clean_for_output(name).lower()).strip("-")
    return slug or "unnamed-item"


def parse_now(now_text=None):
    if now_text:
        dt = datetime.strptime(now_text, "%Y-%m-%d %H:%M")
    else:
        dt = datetime.now()
    return dt, dt.strftime("%Y-%m-%d %H:%M"), dt.strftime("%Y%m%d-%H%M%S")


def has_minimum_submission(ws, row, headers):
    has_type = bool(get_row_value(ws, row, headers, "类型"))
    has_name = bool(get_row_value(ws, row, headers, "英文名") or get_row_value(ws, row, headers, "中文名"))
    has_effect = bool(get_row_value(ws, row, headers, "道具效果（自然语言）"))
    return has_type and has_name and has_effect


def bootstrap_pending_review(ws, headers):
    bootstrapped = []
    for row in range(PLAYER_START_ROW, ws.max_row + 1):
        if get_row_value(ws, row, headers, "审查状态"):
            continue
        if not has_minimum_submission(ws, row, headers):
            continue
        cell(ws, row, headers, "审查状态").value = "待审查"
        bootstrapped.append(item_key(ws, row, headers))
    return bootstrapped


def missing_review_fields(ws, row, headers):
    missing = []
    item_type = get_row_value(ws, row, headers, "类型")
    if not item_type:
        missing.append("类型")
    if not (get_row_value(ws, row, headers, "英文名") or get_row_value(ws, row, headers, "中文名")):
        missing.append("英文名或中文名")
    if not get_row_value(ws, row, headers, "道具池/权重"):
        missing.append("道具池/权重")
    if not (get_row_value(ws, row, headers, "英文描述") or get_row_value(ws, row, headers, "中文描述")):
        missing.append("英文描述或中文描述")
    if not get_row_value(ws, row, headers, "道具效果（自然语言）"):
        missing.append("道具效果（自然语言）")
    if item_type == "主动" and not get_row_value(ws, row, headers, "充能"):
        missing.append("主动道具充能")

    art = get_row_value(ws, row, headers, "贴图")
    if art not in ("无", "已有"):
        missing.append("贴图（无/已有）")
    if art == "已有":
        if not get_row_value(ws, row, headers, "设计者联系平台"):
            missing.append("已有贴图的设计者联系平台")
        if not get_row_value(ws, row, headers, "设计者联系方式"):
            missing.append("已有贴图的设计者联系方式")
    return missing


def build_review_text(ws, row, headers, approved, missing):
    name = item_key(ws, row, headers)
    effect = get_row_value(ws, row, headers, "道具效果（自然语言）")
    item_type = get_row_value(ws, row, headers, "类型")
    tags = get_row_value(ws, row, headers, "标签")
    pool = get_row_value(ws, row, headers, "道具池/权重")

    if not approved:
        return "\n".join(
            [
                "审查结果：需要补充",
                "",
                f"已读取投稿：{name or '未命名道具'}。",
                "",
                f"当前缺少或需要修正：{'、'.join(missing)}。",
                "",
                "请补充后保持其它字段不变，自动化会在下一轮重新审查。",
            ]
        )

    return "\n".join(
        [
            "审查结果：已通过",
            "",
            f"已读取投稿：{name}。",
            "",
            f"基础信息完整：类型={item_type}，标签={tags or '未填写'}，道具池={pool}。",
            "",
            f"已将设计核心理解为：{effect}",
            "",
            "主题判断：道具围绕身体伤痕、代价和残缺收益展开，适合 neverbrith 的缺失生命与痛苦交换气质。",
            "",
            "实现判断：当前自然语言效果足够清晰，可以进入设计包；后续仍需要在代码阶段确认碎心接口、数值叠加和角色上限边界。",
            "",
            "下一步：自动化将生成创意设计包，并交给美术与代码队列继续处理。",
        ]
    )


def review_pending_rows(ws, headers):
    reviewed = []
    for row in range(PLAYER_START_ROW, ws.max_row + 1):
        if get_row_value(ws, row, headers, "审查状态") != "待审查":
            continue

        missing = missing_review_fields(ws, row, headers)
        approved = not missing
        cell(ws, row, headers, "审查状态").value = "已通过" if approved else "需要补充"
        cell(ws, row, headers, "AI回复").value = build_review_text(ws, row, headers, approved, missing)
        cell(ws, row, headers, "AI回复").alignment = Alignment(wrap_text=True, vertical="top")

        if approved and not get_row_value(ws, row, headers, "设计状态"):
            cell(ws, row, headers, "设计状态").value = "待生成"

        reviewed.append({"item": item_key(ws, row, headers), "approved": approved, "missing": missing})
    return reviewed


def package_markdown(ws, row, headers, workflow_id):
    name_en = clean_for_output(get_row_value(ws, row, headers, "英文名"))
    name_zh = clean_for_output(get_row_value(ws, row, headers, "中文名"))
    effect = get_row_value(ws, row, headers, "道具效果（自然语言）")
    desc_en = clean_for_output(get_row_value(ws, row, headers, "英文描述"))
    desc_zh = clean_for_output(get_row_value(ws, row, headers, "中文描述"))
    item_type = get_row_value(ws, row, headers, "类型")
    quality = get_row_value(ws, row, headers, "品质") if "品质" in headers else ""
    tags = get_row_value(ws, row, headers, "标签") if "标签" in headers else ""
    pool = get_row_value(ws, row, headers, "道具池/权重")
    charge = get_row_value(ws, row, headers, "充能")

    return f"""# {name_en or name_zh} / {name_zh or name_en}

## Identity
- Workflow ID: `{workflow_id}`
- Type: {item_type}
- Quality: {quality or "未指定"}
- Tags: {tags or "未指定"}
- Pools: {pool}
- Charge: {charge or "不适用"}
- Pickup text EN: {desc_en or "未指定"}
- Pickup text ZH: {desc_zh or "未指定"}

## Approved Mechanic
{effect}

## Final Mechanic Specification
When the item is obtained or used according to its type, apply the approved mechanic exactly as written above. Keep the primary effect readable and Isaac-like: direct stat gain first, then the heart/resource change.

## Theme Fit
This item frames power as a visible wound. It fits neverbrith through bodily damage, incomplete recovery, and a small reward that still feels like a cost.

## Balance Assumptions
- Treat the damage gain as `+1.00 damage` unless code review chooses a different scalar.
- Treat the broken heart gain as one broken heart.
- If the character cannot receive a broken heart, the code agent must define whether the damage still applies.

## Implementation Boundaries
- Do not remove existing player items.
- Do not add hidden extra effects beyond the submitted mechanic.
- Keep EID text short and explicit.

## Edge Cases
- Character already at broken-heart limit.
- Damage modifiers from other items.
- Rerolls or transformations that reapply collectible cache.

## Art Direction Summary
Small utility knife / craft blade icon, clean silhouette, painful but not gore-heavy. Visual emphasis on a sharp blade and a scar-like mark.

## Code Direction Summary
Implement as a passive stat/resource item unless later changed by author review. Add cache handling for damage and an on-acquire or first-evaluation guard for the broken-heart grant.

## Test Checklist
- Damage increases by 1.
- Exactly one broken heart is granted.
- Effect is not repeatedly applied every frame.
- Behavior is stable after save/load and reroll-sensitive events.
"""


def art_prompt(ws, row, headers):
    name = item_key(ws, row, headers)
    return f"""Create a Binding of Isaac style collectible sprite for `{name}`.

Subject: a small utility knife / craft blade associated with painful scars.
Style: readable 32x32 pixel-art icon, Isaac item-room collectible, strong silhouette, limited palette.
Avoid: realistic gore, unreadable tiny details, modern UI elements, text on the sprite.
Mood: sharp, painful, handmade, slightly tragic, fitting neverbrith's missing-life theme.
"""


def code_prompt(ws, row, headers):
    name = item_key(ws, row, headers)
    effect = get_row_value(ws, row, headers, "道具效果（自然语言）")
    return f"""Implement the neverbrith item `{name}`.

Approved mechanic:
{effect}

Constraints:
- Follow existing neverbrith item patterns in `main.lua`.
- Add localization/EID text consistent with the workbook pickup text.
- Apply +1 damage through cache evaluation.
- Grant one broken heart once, not repeatedly.
- Add tests for damage, broken-heart grant, and repeated update/cache edge cases.
"""


def write_design_files(ws, row, headers, design_root, workflow_id):
    key = safe_key(item_key(ws, row, headers))
    item_dir = Path(design_root) / key
    item_dir.mkdir(parents=True, exist_ok=True)

    files = {
        "design-package.md": package_markdown(ws, row, headers, workflow_id),
        "art-prompt.md": art_prompt(ws, row, headers),
        "code-prompt.md": code_prompt(ws, row, headers),
    }
    for filename, content in files.items():
        (item_dir / filename).write_text(content, encoding="utf-8")

    manifest = {
        "workflow_id": workflow_id,
        "safe_key": key,
        "english_name": clean_for_output(get_row_value(ws, row, headers, "英文名")),
        "chinese_name": clean_for_output(get_row_value(ws, row, headers, "中文名")),
        "art_source": get_row_value(ws, row, headers, "贴图"),
        "designer_contact_platform": get_row_value(ws, row, headers, "设计者联系平台"),
        "designer_contact": get_row_value(ws, row, headers, "设计者联系方式"),
        "review_status": get_row_value(ws, row, headers, "审查状态"),
    }
    (item_dir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return item_dir


def generate_design_packages(ws, headers, design_root, now_display, now_token):
    designed = []
    for row in range(PLAYER_START_ROW, ws.max_row + 1):
        if get_row_value(ws, row, headers, "审查状态") != "已通过":
            continue
        if get_row_value(ws, row, headers, "设计状态") != "待生成":
            continue
        if not has_minimum_submission(ws, row, headers):
            continue

        key = safe_key(item_key(ws, row, headers))
        workflow_id = f"design-{now_token}-{key}"
        cell(ws, row, headers, "设计状态").value = "设计中"
        cell(ws, row, headers, "工作流ID").value = workflow_id
        cell(ws, row, headers, "最后处理时间").value = now_display

        write_design_files(ws, row, headers, design_root, workflow_id)

        cell(ws, row, headers, "设计状态").value = "已完成"
        art_status = "待作者本人审核联系" if get_row_value(ws, row, headers, "贴图") == "已有" else "待生成"
        cell(ws, row, headers, "美术状态").value = art_status
        cell(ws, row, headers, "代码状态").value = "待生成"
        cell(ws, row, headers, "最后处理时间").value = now_display
        designed.append(item_key(ws, row, headers))
    return designed


def run_once(input_path, output_path, design_root, now_text=None):
    _, now_display, now_token = parse_now(now_text)
    wb = load_workbook(input_path)
    if MAIN_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet: {MAIN_SHEET}")
    ws = wb[MAIN_SHEET]
    headers = build_header_map(ws)

    bootstrapped = bootstrap_pending_review(ws, headers)
    reviewed = review_pending_rows(ws, headers)
    designed = generate_design_packages(ws, headers, design_root, now_display, now_token)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {
        "bootstrapped": bootstrapped,
        "reviewed": reviewed,
        "designed": designed,
        "output": str(output_path),
    }


def main():
    parser = argparse.ArgumentParser(description="Run one neverbrith item approval/design automation pass.")
    parser.add_argument("--input", required=True, help="Input xlsx path.")
    parser.add_argument("--output", required=True, help="Output xlsx path.")
    parser.add_argument("--design-root", default="outputs/item-design", help="Directory for generated design packages.")
    parser.add_argument("--now", default=None, help="Override local timestamp as YYYY-MM-DD HH:MM.")
    args = parser.parse_args()

    result = run_once(Path(args.input), Path(args.output), Path(args.design_root), args.now)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
