import importlib.util
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
MODULE_PATH = ROOT / "tools" / "item_automation.py"


def load_module():
    spec = importlib.util.spec_from_file_location("item_automation", MODULE_PATH)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


HEADERS = [
    "序号",
    "类型",
    "英文名",
    "中文名",
    "品质",
    "标签",
    "Cache",
    "充能",
    "初始充能",
    "充能类型",
    "道具池/权重",
    "英文描述",
    "中文描述",
    "道具效果（自然语言）",
    "贴图",
    "设计者联系平台",
    "设计者联系方式",
    "审查状态",
    "AI回复",
    "设计状态",
    "美术状态",
    "代码状态",
    "道具完成度",
    "工作流ID",
    "最后处理时间",
]


def make_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "道具审批"
    ws.append(HEADERS)
    for _ in range(11):
        ws.append([None] * len(HEADERS))
    ws.append(
        [
            13,
            "被动",
            "Utility knife",
            "美工刀",
            2,
            "offensive summonable",
            None,
            None,
            None,
            None,
            "boss(首领房):1",
            "Painful scars",
            "苦痛伤痕",
            "伤害+1，获得一颗碎心",
            "无",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "待处理中",
            None,
            None,
        ]
    )
    ws.append([14, "被动", None, None, None, None, None, None, None, None, None, None, None, None, None])
    ws.append([15, "被动", "Already Approved", "已通过示例", None, None, None, None, None, None, None, None, None, "效果", "无", None, None, "已通过"])
    wb.save(path)


class ItemAutomationTest(unittest.TestCase):
    def test_bootstraps_complete_player_row_without_overwriting_existing_statuses(self):
        module = load_module()
        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.xlsx"
            output_path = Path(tmp) / "output.xlsx"
            design_dir = Path(tmp) / "design"
            make_workbook(input_path)

            result = module.run_once(input_path, output_path, design_dir, now_text="2026-07-02 13:40")

            wb = load_workbook(output_path)
            ws = wb["道具审批"]
            self.assertEqual(ws["R13"].value, "已通过")
            self.assertIn("审查结果：已通过", ws["S13"].value)
            self.assertIsNone(ws["R14"].value)
            self.assertEqual(ws["R15"].value, "已通过")
            self.assertIn("Utility knife", result["bootstrapped"])
            self.assertNotIn("Already Approved", result["bootstrapped"])

    def test_approved_row_generates_design_package_and_advances_downstream_statuses(self):
        module = load_module()
        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.xlsx"
            output_path = Path(tmp) / "output.xlsx"
            design_dir = Path(tmp) / "design"
            make_workbook(input_path)

            result = module.run_once(input_path, output_path, design_dir, now_text="2026-07-02 13:40")

            item_dir = design_dir / "utility-knife"
            self.assertTrue((item_dir / "design-package.md").exists())
            self.assertTrue((item_dir / "art-prompt.md").exists())
            self.assertTrue((item_dir / "code-prompt.md").exists())
            self.assertTrue((item_dir / "manifest.json").exists())
            self.assertEqual(result["designed"], ["Utility knife"])

            wb = load_workbook(output_path)
            ws = wb["道具审批"]
            self.assertEqual(ws["T13"].value, "已完成")
            self.assertEqual(ws["U13"].value, "待生成")
            self.assertEqual(ws["V13"].value, "待生成")
            self.assertRegex(ws["X13"].value, r"^design-20260702-134000-utility-knife$")
            self.assertEqual(ws["Y13"].value, "2026-07-02 13:40")


if __name__ == "__main__":
    unittest.main()
